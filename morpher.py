from openpyxl import load_workbook, Workbook
import pymorphy2

class Phrase:
    WORD_DELIMITERS = [',', ';', ':', ' ']
    MORPHABLE_GRAMMEMES = ['NOUN', 'ADJF', 'ADJS', 'COMP', 'VERB', 'INFN', 'PRTF', 'PRTS', 'NUMR', 'NPRO']

    def __init__(self, phrase_raw):
        self.phrase_raw = phrase_raw
        self.phrase_splitted = self.split_phrase(phrase_raw)

    def split_phrase(self, phrase):
        for delimiter in self.WORD_DELIMITERS:
            phrase = phrase.replace(delimiter, '|' + delimiter + '|')
        phrase = phrase.replace('||', '|')
        return phrase.split('|')

    def morph(self, case, number_category):
        analyzer = pymorphy2.MorphAnalyzer()
        morphed_phrase_items = []
        for item in self.phrase_splitted:
            parsed_item = analyzer.parse(item)[0]
            if parsed_item.tag.POS == 'PREP':
                morphed_phrase_items +=  self.phrase_splitted[self.phrase_splitted.index(item)::]
                break
            if parsed_item.tag.POS in self.MORPHABLE_GRAMMEMES:
                morphed_item = parsed_item.inflect({number_category, case})
                if morphed_item:
                    morphed_phrase_items.append(morphed_item.word)
                else:
                    morphed_phrase_items.append(item)
            else:
                morphed_phrase_items.append(item)
        return ''.join(morphed_phrase_items)

filename = str(input("Original file: "))
sheet_to_read = str(input("Sheet to read: "))
start_column_to_read = str(input("Start column to read (letter): "))
start_row_to_read = int(input("Start row to read (number): "))
start_column_to_save = int(input("Start column to save (number): "))
start_row_to_save = int(input("Start row to save(number): "))

wb = load_workbook(filename)
# ws = wb['Лист1']
ws = wb[sheet_to_read]
column = ws[start_column_to_read]
items_list = [column[x].value for x in range(start_row_to_read,len(column))]

analyzer = pymorphy2.MorphAnalyzer()
morphed_items = []
for item in items_list:
    phrase = Phrase(item)
    morphed_items.append({'им_мн': phrase.morph('nomn', 'plur'),
                          'род_ед': phrase.morph('gent', 'sing'),
                          'род_мн': phrase.morph('gent', 'plur'),
                          'вин_ед': phrase.morph('accs', 'sing'),
                          'предл_мн': phrase.morph('loct', 'plur'),
                          'предл_ед': phrase.morph('loct', 'sing')})

for row_num, item in enumerate(morphed_items, start=start_row_to_save):
    ws.cell(row=row_num, column=start_column_to_save).value = item['им_мн']
    ws.cell(row=row_num, column=start_column_to_save+1).value = item['род_ед']
    ws.cell(row=row_num, column=start_column_to_save+2).value = item['род_мн']
    ws.cell(row=row_num, column=start_column_to_save+3).value = item['вин_ед']
    ws.cell(row=row_num, column=start_column_to_save+4).value = item['предл_мн']
    ws.cell(row=row_num, column=start_column_to_save+5).value = item['предл_ед']

filename_to_save = "morphed_" + filename
wb.save(filename_to_save)
